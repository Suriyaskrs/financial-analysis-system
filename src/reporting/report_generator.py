"""
src/reporting/report_generator.py
------------------------------------
Generates automated financial reports in Excel and CSV formats.
Excel report contains multiple sheets: summary, KPIs, forecasts, anomalies, raw data.
Also generates a text-based executive summary.
"""

import pandas as pd
import numpy as np
import os, sys, datetime, logging
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from src.ingestion.data_loader import (get_financial_records, get_all_companies,
                                        get_forecast_results, get_anomaly_flags,
                                        get_kpi_results)
from src.analytics.kpi_calculator import company_summary, compute_kpis, sector_comparison, latest_kpi_snapshot

logger = logging.getLogger(__name__)

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

# Excel colours
BLUE_DARK   = "1E3A5F"
BLUE_MID    = "2563EB"
BLUE_LIGHT  = "DBEAFE"
GREEN_LIGHT = "D1FAE5"
AMBER_LIGHT = "FEF3C7"
RED_LIGHT   = "FEE2E2"
GREY_LIGHT  = "F8FAFC"
WHITE       = "FFFFFF"


def _style_header_row(ws, row_num: int, fill_hex: str = BLUE_DARK, font_color: str = WHITE):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_color, size=11)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=min_w)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1,
                 header_fill: str = BLUE_DARK, alt_fill: str = GREY_LIGHT):
    """Write a DataFrame to a worksheet with headers and alternating row colours."""
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill  = PatternFill("solid", fgColor=header_fill)
        cell.font  = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    # Data rows
    border = _thin_border()
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        fill_color = alt_fill if (row_idx - start_row) % 2 == 0 else WHITE
        row_fill   = PatternFill("solid", fgColor=fill_color)
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if isinstance(value, (int, float)) else "left")

    _auto_width(ws)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_cover(wb: Workbook, companies_df: pd.DataFrame):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    today = datetime.date.today().strftime("%B %d, %Y")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 60

    title_cell = ws["B2"]
    title_cell.value = "Financial Performance Analytics Report"
    title_cell.font  = Font(bold=True, size=20, color=BLUE_DARK)

    ws["B3"].value = f"Generated: {today}"
    ws["B3"].font  = Font(size=11, color="6B7280")

    ws["B5"].value = "Companies Covered:"
    ws["B5"].font  = Font(bold=True, size=12, color=BLUE_DARK)

    for i, (_, row) in enumerate(companies_df.iterrows(), 6):
        ws.cell(row=i, column=2).value = f"  •  {row['company_name']} ({row['sector']}, {row['country']})"
        ws.cell(row=i, column=2).font  = Font(size=11)

    ws["B12"].value = "Contents:"
    ws["B12"].font  = Font(bold=True, size=12, color=BLUE_DARK)
    sheets = ["Executive Summary", "Company KPIs", "Sector Analysis",
              "Forecasts", "Anomaly Flags", "Raw Financial Data"]
    for i, s in enumerate(sheets, 13):
        ws.cell(row=i, column=2).value = f"  {i-12}. {s}"
        ws.cell(row=i, column=2).font  = Font(size=11)


def _sheet_executive_summary(wb: Workbook, df_rec, companies_df):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20

    snap = latest_kpi_snapshot(df_rec, companies_df)
    summary = company_summary(df_rec, companies_df)

    ws["B2"].value = "Executive Summary — Latest Quarter"
    ws["B2"].font  = Font(bold=True, size=16, color=BLUE_DARK)

    # KPI cards (top 3 performers)
    ws["B4"].value = "Top Performer by Profit Margin"
    ws["B4"].font  = Font(bold=True, size=11)
    best = summary.iloc[0]
    ws["B5"].value = f"{best['company_name']} ({best['sector']})"
    ws["B6"].value = f"Profit Margin: {best['avg_profit_margin']:.1f}%"
    ws["B6"].font  = Font(color="059669", bold=True)

    ws["B8"].value = "Highest Revenue"
    ws["B8"].font  = Font(bold=True, size=11)
    top_rev = summary.sort_values("total_revenue", ascending=False).iloc[0]
    ws["B9"].value  = top_rev["company_name"]
    ws["B10"].value = f"£{top_rev['total_revenue']/1000:.1f}M total (3-year)"
    ws["B10"].font  = Font(color=BLUE_MID, bold=True)

    ws["B12"].value = "Summary Table"
    ws["B12"].font  = Font(bold=True, size=13, color=BLUE_DARK)

    display_cols = ["company_name","sector","avg_profit_margin","avg_expense_ratio",
                    "avg_roa","avg_revenue_growth","total_revenue"]
    _df_to_sheet(ws, summary[display_cols].rename(columns={
        "company_name":      "Company",
        "sector":            "Sector",
        "avg_profit_margin": "Avg Profit Margin %",
        "avg_expense_ratio": "Avg Expense Ratio %",
        "avg_roa":           "Avg ROA %",
        "avg_revenue_growth":"Avg Revenue Growth %",
        "total_revenue":     "Total Revenue (£k)",
    }), start_row=13)


def _sheet_kpis(wb: Workbook, df_rec, companies_df):
    ws = wb.create_sheet("Company KPIs")
    df  = compute_kpis(df_rec)
    df  = df.merge(companies_df[["company_id","company_name"]], on="company_id")

    disp = df[["company_name","date","revenue","profit","profit_margin",
               "operating_ratio","expense_ratio","return_on_assets","revenue_growth"]].copy()
    disp["date"]           = disp["date"].astype(str)
    disp["profit_margin"]  = (disp["profit_margin"]  * 100).round(2)
    disp["operating_ratio"]= (disp["operating_ratio"] * 100).round(2)
    disp["expense_ratio"]  = (disp["expense_ratio"]   * 100).round(2)
    disp["return_on_assets"]=(disp["return_on_assets"]* 100).round(2)
    disp["revenue_growth"] = (disp["revenue_growth"]  * 100).round(2)

    disp.columns = ["Company","Date","Revenue (£k)","Profit (£k)",
                    "Profit Margin %","Operating Ratio %","Expense Ratio %",
                    "ROA %","Revenue Growth %"]
    _df_to_sheet(ws, disp)


def _sheet_sector(wb: Workbook, df_rec, companies_df):
    ws = wb.create_sheet("Sector Analysis")
    sec = sector_comparison(df_rec, companies_df)
    sec["total_revenue"] = sec["total_revenue"].round(0)
    sec.columns = ["Sector","Avg Profit Margin %","Avg Expense Ratio %",
                   "Avg ROA %","Total Revenue (£k)","# Companies"]
    _df_to_sheet(ws, sec)


def _sheet_forecasts(wb: Workbook, forecasts_df: pd.DataFrame, companies_df: pd.DataFrame):
    ws = wb.create_sheet("Forecasts")
    if forecasts_df.empty:
        ws["A1"].value = "No forecast data. Run forecasts first."
        return

    df = forecasts_df.merge(companies_df[["company_id","company_name"]], on="company_id")
    disp = df[["company_name","forecast_date","predicted_revenue","predicted_profit",
               "lower_bound","upper_bound","model_used","mae","mape"]].copy()
    disp["forecast_date"] = disp["forecast_date"].astype(str)
    disp.columns = ["Company","Forecast Date","Predicted Revenue","Predicted Profit",
                    "Lower Bound","Upper Bound","Model","MAE","MAPE %"]
    _df_to_sheet(ws, disp)


def _sheet_anomalies(wb: Workbook, anomalies_df: pd.DataFrame, companies_df: pd.DataFrame):
    ws = wb.create_sheet("Anomaly Flags")
    if anomalies_df.empty:
        ws["A1"].value = "No anomalies detected."
        return

    df = anomalies_df.merge(companies_df[["company_id","company_name"]], on="company_id")
    disp = df[["company_name","date","metric","value","z_score","severity"]].copy()
    disp["date"] = disp["date"].astype(str)
    disp.columns = ["Company","Date","Metric","Value","Z-Score","Severity"]
    _df_to_sheet(ws, disp)

    # Colour-code severity
    sev_col = 6
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=sev_col, max_col=sev_col):
        for cell in row:
            if cell.value == "SEVERE":
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
                cell.font = Font(bold=True, color="DC2626")
            elif cell.value == "WARNING":
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
                cell.font = Font(color="D97706")


def _sheet_raw_data(wb: Workbook, df_rec: pd.DataFrame, companies_df: pd.DataFrame):
    ws = wb.create_sheet("Raw Financial Data")
    df = df_rec.merge(companies_df[["company_id","company_name","sector"]], on="company_id")
    disp = df[["company_name","sector","date","revenue","expenses","profit",
               "operating_cost","assets","liabilities"]].copy()
    disp["date"] = disp["date"].astype(str)
    disp.columns = ["Company","Sector","Date","Revenue","Expenses","Profit",
                    "Operating Cost","Assets","Liabilities"]
    _df_to_sheet(ws, disp)


# ── Main Report Generator ─────────────────────────────────────────────────────

def generate_excel_report(filename: str = None) -> str:
    """Generate full multi-sheet Excel report. Returns file path."""
    filename = filename or f"financial_report_{datetime.date.today()}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    df_rec       = get_financial_records()
    df_rec["date"] = pd.to_datetime(df_rec["date"])
    companies_df = get_all_companies()
    forecasts_df = get_forecast_results()
    anomalies_df = get_anomaly_flags()

    wb = Workbook()
    _sheet_cover(wb, companies_df)
    _sheet_executive_summary(wb, df_rec, companies_df)
    _sheet_kpis(wb, df_rec, companies_df)
    _sheet_sector(wb, df_rec, companies_df)
    _sheet_forecasts(wb, forecasts_df, companies_df)
    _sheet_anomalies(wb, anomalies_df, companies_df)
    _sheet_raw_data(wb, df_rec, companies_df)

    wb.save(filepath)
    logger.info(f"Excel report saved: {filepath}")
    return filepath


def generate_csv_report(filename: str = None) -> str:
    """Export processed KPI data as a single CSV for further analysis."""
    filename = filename or f"kpi_report_{datetime.date.today()}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)

    df_rec       = get_financial_records()
    df_rec["date"] = pd.to_datetime(df_rec["date"])
    companies_df = get_all_companies()
    df_kpi       = compute_kpis(df_rec)
    df_kpi       = df_kpi.merge(companies_df, on="company_id")
    df_kpi["date"] = df_kpi["date"].astype(str)

    df_kpi.to_csv(filepath, index=False)
    logger.info(f"CSV report saved: {filepath}")
    return filepath


def generate_all_reports() -> dict:
    """Generate both Excel and CSV reports. Returns dict of file paths."""
    excel_path = generate_excel_report()
    csv_path   = generate_csv_report()
    return {"excel": excel_path, "csv": csv_path}
